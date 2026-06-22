import streamlit as st
import requests
import base64
import json
import io
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Config ────────────────────────────────────
st.set_page_config(page_title="Form OCR → Excel", page_icon="📋", layout="centered")
st.title("📋 Handwritten Form → Excel")
st.caption("Image do → LLaVA padhega → Excel download karo!")

OLLAMA_URL = "http://localhost:11434/api/generate"

FIELDS = [
    "Full Name",
    "Phone Number",
    "Address",
    "City, State, Zip",
    "Email",
    "Date of Birth",
    "SSN",
    "Marital Status",
    "Bank Name",
    "Account Type",
    "Routing Number",
    "Account Number",
]

# ── Ollama status ─────────────────────────────
def check_ollama():
    try:
        r = requests.get("http://localhost:11434/api/tags", timeout=2)
        models = [m["name"] for m in r.json().get("models", [])]
        llava_models = [m for m in models if "llava" in m.lower()]
        return True, llava_models
    except:
        return False, []

# ── Image → Base64 ────────────────────────────
def image_to_base64(image: Image.Image) -> str:
    buf = io.BytesIO()
    image.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")

# ── LLaVA call ───────────────────────────────
def extract_with_llava(image: Image.Image, model_name: str) -> dict:
    img_b64 = image_to_base64(image)

    prompt = f"""This is a handwritten employee form image.
Extract ONLY the handwritten/filled values for these exact fields:
{json.dumps(FIELDS, indent=2)}

Rules:
- Printed labels on LEFT side = field names (ignore them)
- Handwritten text on RIGHT side = values (extract these)
- If a field is empty or not visible, use empty string ""
- Return ONLY valid JSON, no explanation, no markdown

Output format:
{{
  "Full Name": "...",
  "Phone Number": "...",
  "Address": "...",
  "City, State, Zip": "...",
  "Email": "...",
  "Date of Birth": "...",
  "SSN": "...",
  "Marital Status": "...",
  "Bank Name": "...",
  "Account Type": "...",
  "Routing Number": "...",
  "Account Number": "..."
}}"""

    response = requests.post(
        OLLAMA_URL,
        json={
            "model": model_name,
            "prompt": prompt,
            "images": [img_b64],
            "stream": False,
            "format": "json",
            "options": {"temperature": 0.1},
        },
        timeout=120,
    )
    response.raise_for_status()
    raw = response.json().get("response", "").strip()
    raw = raw.replace("```json", "").replace("```", "").strip()
    result = json.loads(raw)

    # Ensure all fields exist
    for f in FIELDS:
        if f not in result:
            result[f] = ""

    return result

# ── Excel builder ─────────────────────────────
def build_excel(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    h_fill = PatternFill("solid", fgColor="1A3A5C")
    k_fill = PatternFill("solid", fgColor="D6EAF8")
    v_fill = PatternFill("solid", fgColor="FDFEFE")
    thin   = Side(style="thin", color="AAAAAA")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["A1"] = "Field"
    ws["B1"] = "Handwritten Value"
    for cell in [ws["A1"], ws["B1"]]:
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = center
        cell.border = brd
    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45

    for i, (field, value) in enumerate(data.items(), start=2):
        ka = ws.cell(row=i, column=1, value=field)
        va = ws.cell(row=i, column=2, value=value)
        ka.fill = k_fill
        va.fill = v_fill
        for c in [ka, va]:
            c.border = brd
            c.alignment = left
            c.font = Font(name="Arial", size=11)
        ws.row_dimensions[i].height = 22

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── UI ────────────────────────────────────────
ollama_ok, llava_models = check_ollama()

if not ollama_ok:
    st.error("🔴 Ollama nahi chal raha! Terminal mein `ollama serve` chala ke aao.")
    st.stop()

if not llava_models:
    st.error("🔴 LLaVA model nahi mila! Terminal mein chalao:")
    st.code("ollama pull llava", language="bash")
    st.stop()

st.success(f"🟢 Ollama ready! LLaVA model: `{llava_models[0]}`")

# Model selector (agar multiple llava versions hain)
selected_model = st.selectbox("LLaVA Model", llava_models) if len(llava_models) > 1 else llava_models[0]

uploaded = st.file_uploader("📤 Form Image Upload Karo", type=["jpg", "jpeg", "png"])

if uploaded:
    image = Image.open(uploaded)
    st.image(image, caption="Uploaded Form", use_container_width=True)

    if st.button("🔍 Extract Karo", type="primary", use_container_width=True):
        with st.spinner("LLaVA image padh raha hai... 🧠 (30-60 sec)"):
            try:
                data = extract_with_llava(image, selected_model)

                st.divider()
                st.subheader("✅ Extracted Values")

                # Show table
                table = [{"Field": k, "Value": v} for k, v in data.items()]
                st.dataframe(
                    table,
                    use_container_width=True,
                    column_config={
                        "Field": st.column_config.TextColumn("🔑 Field", width="medium"),
                        "Value": st.column_config.TextColumn("✍️ Handwritten Value", width="large"),
                    },
                    hide_index=True,
                )

                # Manual edit
                st.subheader("✏️ Galat ho toh yahan fix karo")
                edited = {}
                c1, c2 = st.columns(2)
                for i, (field, value) in enumerate(data.items()):
                    with (c1 if i % 2 == 0 else c2):
                        edited[field] = st.text_input(field, value=value, key=f"e{i}")

                # Download
                excel_bytes = build_excel(edited)
                st.download_button(
                    "📥 Excel Download Karo",
                    data=excel_bytes,
                    file_name="form_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"Error aaya: {e}")
                st.info("LLaVA model sahi se pull hua hai? `ollama pull llava` try karo.")
