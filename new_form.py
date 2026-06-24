import streamlit as st
import requests
import base64
import json
import io
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Config ────────────────────────────────────
st.set_page_config(page_title="HAL Form OCR", page_icon="📋", layout="centered")
st.title("📋 HAL Application Form → Excel")
st.caption("Image upload karo → LLaVA extract karega → Edit karo → Excel download!")

OLLAMA_URL = "http://localhost:11434/api/generate"

# ── HAL Form Fields ───────────────────────────
FIELDS = [
    "Trade Applied For",
    "Full Name",
    "Present Local Address",
    "Permanent Home Address",
    "Father's Full Name",
    "Father's Present Postal Address",
    "Father's Permanent Home Address",
    "Father's Profession",
    "Father's Designation & Official Address",
    "Date of Birth",
    "Present Age",
    "Nationality",
    "Religion",
    "Email",
    "Mobile Number",
    "SC/ST Member (Yes/No)",
    "Married or Single",
    "School/College 1 - Name & Address",
    "School/College 1 - Date of Entering",
    "School/College 1 - Date of Leaving",
    "School/College 1 - Examination Passed",
    "School/College 2 - Name & Address",
    "School/College 2 - Date of Entering",
    "School/College 2 - Date of Leaving",
    "School/College 2 - Examination Passed",
    "School/College 3 - Name & Address",
    "School/College 3 - Date of Entering",
    "School/College 3 - Date of Leaving",
    "School/College 3 - Examination Passed",
]

# ── Ollama check ──────────────────────────────
def check_ollama():
    try:
        r = requests.get("http://localhost:11434/api/tags", timeout=2)
        models = [m["name"] for m in r.json().get("models", [])]
        llava  = [m for m in models if "llava" in m.lower()]
        return True, llava
    except:
        return False, []

# ── Image → Base64 ────────────────────────────
def to_b64(image: Image.Image) -> str:
    buf = io.BytesIO()
    image.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")

# ── LLaVA Extract ─────────────────────────────
def extract_with_llava(image: Image.Image, model: str) -> dict:
    prompt = """This is a handwritten application form (HAL - Hindustan Aeronautics Limited).
Extract ALL filled/handwritten values from this form.

Important rules:
- Printed labels = field names (ignore)
- Handwritten/filled text = values (extract these)
- Grid boxes: read letters together as one word (e.g. S-H-A-L-I-N-E-E = SHALINEE)
- Multi-line fields: join all lines with space
- If field is empty → use ""
- For school/college table: extract each row separately

Return ONLY this exact JSON (no explanation, no markdown):
{
  "Trade Applied For": "",
  "Full Name": "",
  "Present Local Address": "",
  "Permanent Home Address": "",
  "Father's Full Name": "",
  "Father's Present Postal Address": "",
  "Father's Permanent Home Address": "",
  "Father's Profession": "",
  "Father's Designation & Official Address": "",
  "Date of Birth": "",
  "Present Age": "",
  "Nationality": "",
  "Religion": "",
  "Email": "",
  "Mobile Number": "",
  "SC/ST Member (Yes/No)": "",
  "Married or Single": "",
  "School/College 1 - Name & Address": "",
  "School/College 1 - Date of Entering": "",
  "School/College 1 - Date of Leaving": "",
  "School/College 1 - Examination Passed": "",
  "School/College 2 - Name & Address": "",
  "School/College 2 - Date of Entering": "",
  "School/College 2 - Date of Leaving": "",
  "School/College 2 - Examination Passed": "",
  "School/College 3 - Name & Address": "",
  "School/College 3 - Date of Entering": "",
  "School/College 3 - Date of Leaving": "",
  "School/College 3 - Examination Passed": ""
}"""

    resp = requests.post(
        OLLAMA_URL,
        json={
            "model": model,
            "prompt": prompt,
            "images": [to_b64(image)],
            "stream": False,
            "format": "json",
            "options": {"temperature": 0.1},
        },
        timeout=180,
    )
    resp.raise_for_status()
    raw = resp.json().get("response", "").strip()
    raw = raw.replace("```json", "").replace("```", "").strip()

    try:
        result = json.loads(raw)
    except:
        # fallback: empty dict
        result = {}

    # ensure all keys exist
    for f in FIELDS:
        if f not in result:
            result[f] = ""

    return result

# ── Excel builder ─────────────────────────────
def build_excel(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HAL Form Data"

    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    h_fill = PatternFill("solid", fgColor="1A3A5C")
    k_fill = PatternFill("solid", fgColor="D6EAF8")
    v_fill = PatternFill("solid", fgColor="FDFEFE")
    thin   = Side(style="thin", color="AAAAAA")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["A1"] = "Field"
    ws["B1"] = "Extracted Value"
    for cell in [ws["A1"], ws["B1"]]:
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = center
        cell.border    = brd
    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50

    for i, (field, value) in enumerate(data.items(), start=2):
        ka = ws.cell(row=i, column=1, value=field)
        va = ws.cell(row=i, column=2, value=value)
        ka.fill = k_fill
        va.fill = v_fill
        for c in [ka, va]:
            c.border    = brd
            c.alignment = left
            c.font      = Font(name="Arial", size=11)
        ws.row_dimensions[i].height = 22

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── UI ────────────────────────────────────────
ollama_ok, llava_models = check_ollama()

if not ollama_ok:
    st.error("🔴 Ollama nahi chal raha! `ollama serve` chala ke aao.")
    st.stop()

if not llava_models:
    st.error("🔴 LLaVA model nahi mila!")
    st.code("ollama pull llava", language="bash")
    st.stop()

st.success(f"🟢 Ollama ready! Model: `{llava_models[0]}`")
selected_model = (
    st.selectbox("LLaVA Model choose karo", llava_models)
    if len(llava_models) > 1
    else llava_models[0]
)

st.divider()
uploaded = st.file_uploader("📤 Form Image Upload Karo (JPG/PNG)", type=["jpg","jpeg","png"])

if uploaded:
    image = Image.open(uploaded)
    st.image(image, caption="Uploaded Form", use_container_width=True)

    if st.button("🔍 Extract Karo", type="primary", use_container_width=True):

        with st.spinner("LLaVA padh raha hai... 🧠 (1-2 min)"):
            try:
                data = extract_with_llava(image, selected_model)
                st.session_state["extracted"] = data
                st.session_state["edited"]    = dict(data)
            except Exception as e:
                st.error(f"Error: {e}")

# ── Show result + edit + download ────────────
if "extracted" in st.session_state:
    st.divider()
    st.subheader("✅ Extracted Values")

    # Preview table
    table = [{"Field": k, "Value": v} for k, v in st.session_state["extracted"].items()]
    st.dataframe(
        table,
        use_container_width=True,
        column_config={
            "Field": st.column_config.TextColumn("🔑 Field",  width="large"),
            "Value": st.column_config.TextColumn("✍️ Value", width="large"),
        },
        hide_index=True,
    )

    st.divider()
    st.subheader("✏️ Galat ho toh yahan fix karo")

    edited = {}

    # Group fields for cleaner UI
    groups = {
        "👤 Personal Details": FIELDS[0:2],
        "🏠 Address": FIELDS[2:4],
        "👨 Father's Details": FIELDS[4:9],
        "📅 DOB & Age": FIELDS[9:11],
        "🌍 Other Details": FIELDS[11:17],
        "🏫 School / College 1": FIELDS[17:21],
        "🏫 School / College 2": FIELDS[21:25],
        "🏫 School / College 3": FIELDS[25:29],
    }

    for group_name, group_fields in groups.items():
        with st.expander(group_name, expanded=True):
            for field in group_fields:
                edited[field] = st.text_input(
                    field,
                    value=st.session_state["edited"].get(field, ""),
                    key=f"edit_{field}"
                )

    st.session_state["edited"] = edited

    st.divider()
    excel_bytes = build_excel(edited)
    st.download_button(
        "📥 Excel Download Karo",
        data=excel_bytes,
        file_name="HAL_form_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
