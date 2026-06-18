import streamlit as st
import easyocr
import io
from PIL import Image
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from transformers import TrOCRProcessor, VisionEncoderDecoderModel

st.set_page_config(page_title="Form OCR → Excel", page_icon="📋", layout="wide")
st.title("📋 Form OCR → Excel")
st.caption("Fixed fields (left) + Handwritten values (right) → Excel download")

# ── Fixed Fields ──────────────────────────────
FIELDS = [
    "Full Name",
    "Phone Number",
    "Address",
    "City, State, Zip",
    "Email",
    "Date of Birth",
    "SSN",
    "Marital Status",
    "Bank Name",
    "Account Type",
    "Routing Number",
    "Account Number",
]

# Words that indicate a row is a section header / title — SKIP these
SKIP_KEYWORDS = [
    "employee information sheet",
    "employee details", "confidential details", "banking details",
    "personal details", "contact details",
]

# ── Models ────────────────────────────────────
@st.cache_resource(show_spinner="Loading EasyOCR...")
def load_easyocr():
    return easyocr.Reader(["en"], gpu=False)

@st.cache_resource(show_spinner="Loading TrOCR (first run ~1 min)...")
def load_trocr():
    processor = TrOCRProcessor.from_pretrained("microsoft/trocr-base-handwritten")
    model = VisionEncoderDecoderModel.from_pretrained("microsoft/trocr-base-handwritten")
    return processor, model

def trocr_read(crop_pil, processor, model):
    pixel_values = processor(images=crop_pil.convert("RGB"), return_tensors="pt").pixel_values
    ids = model.generate(pixel_values)
    return processor.batch_decode(ids, skip_special_tokens=True)[0].strip()

# ── Main extraction ───────────────────────────
def extract_values(image: Image.Image, use_trocr: bool, conf_thresh: float, split_pct: float):
    img_array = np.array(image.convert("RGB"))
    H, W = img_array.shape[:2]

    reader = load_easyocr()
    results = reader.readtext(img_array, detail=1)
    results = [r for r in results if r[2] >= conf_thresh]

    split_x   = W * split_pct
    top_skip  = H * 0.08          # top 8% = title area, skip it

    right_items = []
    debug_skipped = []

    for (bbox, text, conf) in results:
        pts      = np.array(bbox)
        x_center = pts[:, 0].mean()
        y_center = pts[:, 1].mean()

        # Skip title area at the very top
        if y_center < top_skip:
            debug_skipped.append(f"[TOP SKIP] {text}")
            continue

        # Skip section headers anywhere in the image
        if text.strip().lower().rstrip(":") in SKIP_KEYWORDS:
            debug_skipped.append(f"[HEADER SKIP] {text}")
            continue

        # Only take RIGHT side items
        if x_center < split_x:
            continue

        right_items.append({
            "bbox":    bbox,
            "text":    text,
            "conf":    conf,
            "y_center": y_center,
            "x_min":  int(pts[:, 0].min()),
            "x_max":  int(pts[:, 0].max()),
            "y_min":  int(pts[:, 1].min()),
            "y_max":  int(pts[:, 1].max()),
        })

    # Sort top→bottom
    right_items.sort(key=lambda x: x["y_center"])

    # Merge items on same row (y within 18px)
    merged_rows = []
    used = set()
    for i, item in enumerate(right_items):
        if i in used:
            continue
        group = [item]
        used.add(i)
        for j, other in enumerate(right_items):
            if j in used:
                continue
            if abs(other["y_center"] - item["y_center"]) < 18:
                group.append(other)
                used.add(j)
        group.sort(key=lambda x: x["x_min"])
        merged_rows.append({
            "text":        " ".join(g["text"] for g in group),
            "y_center":    np.mean([g["y_center"] for g in group]),
            "crop_coords": (
                min(g["x_min"] for g in group),
                min(g["y_min"] for g in group),
                max(g["x_max"] for g in group),
                max(g["y_max"] for g in group),
            ),
        })

    merged_rows.sort(key=lambda x: x["y_center"])

    # Skip rows whose merged text looks like a section header
    filtered_rows = []
    for row in merged_rows:
        if row["text"].strip().lower().rstrip(":") in SKIP_KEYWORDS:
            debug_skipped.append(f"[MERGED HEADER SKIP] {row['text']}")
            continue
        filtered_rows.append(row)

    # Re-read with TrOCR if enabled
    if use_trocr:
        processor, model = load_trocr()

    values = []
    for row in filtered_rows:
        if use_trocr:
            x1, y1, x2, y2 = row["crop_coords"]
            pad = 5
            x1 = max(0, x1-pad); y1 = max(0, y1-pad)
            x2 = min(W, x2+pad); y2 = min(H, y2+pad)
            crop_pil = Image.fromarray(img_array[y1:y2, x1:x2])
            text = trocr_read(crop_pil, processor, model)
        else:
            text = row["text"]
        values.append(text)

    # Map FIELDS → values 1-to-1
    result = {field: (values[i] if i < len(values) else "") for i, field in enumerate(FIELDS)}
    return result, filtered_rows, debug_skipped

# ── Excel builder ─────────────────────────────
def build_excel(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    h_fill = PatternFill("solid", fgColor="1A3A5C")
    k_fill = PatternFill("solid", fgColor="D6EAF8")
    v_fill = PatternFill("solid", fgColor="FDFEFE")
    thin   = Side(style="thin", color="AAAAAA")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["A1"] = "Field"
    ws["B1"] = "Handwritten Value"
    for cell in [ws["A1"], ws["B1"]]:
        cell.font=h_font; cell.fill=h_fill; cell.alignment=center; cell.border=brd
    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45

    for i, (field, value) in enumerate(data.items(), start=2):
        ka = ws.cell(row=i, column=1, value=field)
        va = ws.cell(row=i, column=2, value=value)
        ka.fill=k_fill; va.fill=v_fill
        for c in [ka, va]:
            c.border=brd; c.alignment=left; c.font=Font(name="Arial", size=11)
        ws.row_dimensions[i].height = 22

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()

# ── UI ────────────────────────────────────────
col1, col2 = st.columns([1, 1], gap="large")

with col1:
    st.subheader("📤 Upload Form Image")
    uploaded    = st.file_uploader("JPG / PNG", type=["jpg","jpeg","png"])
    use_trocr   = st.toggle("TrOCR use karo (accurate, slow)", value=True)
    conf_thresh = st.slider("EasyOCR confidence", 0.1, 0.9, 0.30, 0.05)
    split_pct   = st.slider(
        "Left/Right split % (adjust agar values shift ho rahi hain)",
        0.20, 0.60, 0.40, 0.01,
        help="Image width ka kitna % left side = labels. Baaki = handwritten values."
    )
    run_btn = st.button("🔍 Extract", type="primary", use_container_width=True)
    if uploaded:
        st.image(Image.open(uploaded), use_container_width=True)

with col2:
    st.subheader("🔑 Fixed Fields")
    for f in FIELDS:
        st.markdown(f"• **{f}**")

if run_btn and uploaded:
    image = Image.open(uploaded)
    with st.spinner("Extracting... ⏳"):
        data, rows, skipped = extract_values(image, use_trocr, conf_thresh, split_pct)

    st.divider()

    # Debug expander
    with st.expander("🔍 Debug — kya skip hua / kya detect hua"):
        st.markdown("**Skipped rows:**")
        st.code("\n".join(skipped) if skipped else "Nothing skipped", language=None)
        st.markdown("**Right-side rows detected (in order):**")
        st.code("\n".join(r["text"] for r in rows), language=None)

    st.subheader("✅ Result")
    table_data = [{"Field": k, "Handwritten Value": v} for k, v in data.items()]
    st.dataframe(
        table_data,
        use_container_width=True,
        column_config={
            "Field":             st.column_config.TextColumn("🔑 Field", width="medium"),
            "Handwritten Value": st.column_config.TextColumn("✍️ Value", width="large"),
        },
        hide_index=True,
    )

    st.subheader("✏️ Fix karo agar kuch galat ho")
    edited = {}
    cols_e = st.columns(2)
    for i, (field, value) in enumerate(data.items()):
        with cols_e[i % 2]:
            edited[field] = st.text_input(field, value=value, key=f"e{i}")

    excel_bytes = build_excel(edited)
    st.download_button(
        "📥 Download Excel",
        data=excel_bytes,
        file_name="form_extracted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

elif run_btn:
    st.error("Pehle image upload karo! 😄")
